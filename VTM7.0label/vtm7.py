import pandas as pd 
import os
from openpyxl import Workbook,load_workbook
import re
import numpy as np


path = r'./vtm7.xlsx'#这个是所有excel文件夹的路径
data_list=[] 
Depth_list=[]
f=open(path,'rb')
data=pd.read_excel(f)

x=data.shape

length=x[0]
data_list=list(data.iloc[0:length-1,0])
pattern =re.compile(r'\d+')

Depth_list=[]
for data_temp in data_list:
    if((data_temp.find('SplitSeries')!=-1)and(data_temp.find('POC')!=-1)and(data_temp.find('Chroma')==-1)):
        Depth_list.append(data_temp)
rows=len(Depth_list)
cols=6
rowx=0
colx=0


BlockDataStore=np.zeros([rows,cols])
DataTempStore=[]
rowx=0
for Depth_temp in Depth_list:
    Data_temp=pattern.findall(Depth_temp)

    colx=0
    for Data in Data_temp:
        BlockDataStore[rowx][colx]=Data
        colx=colx+1

    rowx=rowx+1
Dim=BlockDataStore.shape
Dec_Temp=0
Bin_Temp=[]
GroundTruth=['1']
for rowx in range(Dim[0]):
    Dec_Temp=BlockDataStore[rowx][-1]
    Bin_Temp=list(bin(int(Dec_Temp)))
    Bin_Temp.reverse()
    count=0
    Bin_Len=len(Bin_Temp)
    count_zero=0
    for count in range(Bin_Len):
        ##因列表反向，相当于从低位到高位遍历
        if(count==Bin_Len-2):
            GroundTruth.append('0')
            BlockDataStore[rowx][-1]=int(("".join(GroundTruth)))##list转字符串
            GroundTruth=['1']
            break
        elif(Bin_Temp[count]=='1'):
            ##m每四个0为一个有效元素
            if((count_zero==1)or(count_zero==0)):
                continue
            elif(count_zero==4):
                #两个连续的1
                if(Bin_Temp[count+1]=='1'):##11
                    GroundTruth.append('3')
                elif((Bin_Temp[count+1]=='0')and(Bin_Temp[count+2]=='1')):##101
                    GroundTruth.append('5')
                else:#1
                    GroundTruth.append('1')
            elif(count_zero==5):
                ##10
                GroundTruth.append('2')
            elif(count_zero==6):
                #100
                GroundTruth.append('4')
            count_zero=0
            # print(GroundTruth)
        else:
            count_zero=count_zero+1
            
poc_sum=5
Dim=BlockDataStore.shape
count0=0
count1=0
count2=0
count3=0
count4=0
for rowx in range(Dim[0]):
    if(BlockDataStore[rowx][0]==0):
        count0=count0+1
    elif(BlockDataStore[rowx][0]==1):
        count1=count1+1
    elif(BlockDataStore[rowx][0]==2):
        count2=count2+1
    elif(BlockDataStore[rowx][0]==3):
        count3=count3+1
    elif(BlockDataStore[rowx][0]==4):
        count4=count4+1

count_List=[]
count_List.append(count0)
count_List.append(count1)
count_List.append(count2)
count_List.append(count3)
count_List.append(count4)

max(count_List)
count_List

poc_sum=5
Dim=BlockDataStore.shape
count0=0
count1=0
count2=0
count3=0
count4=0
count_temp=0
BlockDataStoreinFrame=np.zeros([max(count_List),6,5])
for index in range(len(count_List)):
    print(BlockDataStoreinFrame[0:count_List[index],:,index].shape)
    BlockDataStoreinFrame[0:count_List[index],:,index]=BlockDataStore[count_temp:(count_temp+count_List[index]),:]
    count_temp=count_temp+count_List[index]


Dim3=BlockDataStoreinFrame.shape
for ii in range(0,poc_sum):
    wb = Workbook()
    wb.save('./GroundTruthoutput'+str(ii)+'.xlsx')
for poc_index in range(Dim3[-1]):
    filepath_write = './GroundTruthoutput'+str(int(poc_index))+'.xlsx'
    wb = load_workbook(filepath_write)
    ws1 = ws1 = wb.create_sheet(title="Sheet")
    row_ind=1
    col_ind=1
    for rowx in range(Dim3[0]):
        row_xlsx=int(BlockDataStoreinFrame[rowx,1,poc_index]+1)
        col_xlsx=int(BlockDataStoreinFrame[rowx,2,poc_index]+1)
        row_xlsx_sum=int(BlockDataStoreinFrame[rowx,3,poc_index])
        col_xlsx_sum=int(BlockDataStoreinFrame[rowx,4,poc_index])

        for row_index in range(row_xlsx,row_xlsx+row_xlsx_sum):
            for col_index in range(col_xlsx,col_xlsx+col_xlsx_sum):
                ws1.cell(row=row_index,column=col_index).value=BlockDataStoreinFrame[rowx,-1,poc_index]
    wb.save(filepath_write)
