import pandas as pd 
import os
from openpyxl import Workbook,load_workbook
path = r'/Users/kobezhe/Downloads/file'#这个是所有excel文件夹的路径
pathfile = '/Users/kobezhe/Downloads/file/'#这个比上面那个多了个/，修改的化改这两个就行
filenames = os.listdir(path)
print(filenames)
data1=[]
filenames.remove('.DS_Store')#移除备份文件
#print(filenames)

for file in filenames:
    #matchObj = re.search(r'DS_Store',file)
    if os.path.isfile(pathfile+file):
        filepath = pathfile +file
        print(filepath)
    else:
        path1 = pathfile + file 
        filenames_next = os.listdir(path1)
        for file_next in filenames_next:
            filepath = path1 + '/' + file_next

    f= open(filepath,'rb')
    data = pd.read_excel(f)

    data1.append(list(data.iloc[2,1:7].values))

filepath_write = '/Users/kobezhe/Downloads/file/1602大班2018-2019学年综合量化汇总表.xlsx'
wb = load_workbook(filepath_write)
ws1 = wb['Sheet1']
col_len = len(filenames)
row_index = 0
col_index = 0
for rowx in range(4,4+col_len):
    col_index = 0    
    for colx in range(2,7):
        ws1.cell(row=rowx,column=colx).value=data1[row_index][col_index]
        col_index=col_index+1
    row_index = row_index+1

wb.save(filepath_write)